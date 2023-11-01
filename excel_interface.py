import pylogix
from PyQt5 import Qt
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from PyQt5.QtWidgets import QMessageBox, QProgressDialog, QDialog, QCheckBox
from PyQt5.QtCore import Qt, QVariant

import xlsxwriter
from xlsxwriter import exceptions as xlsx_exc
import openpyxl
import openpyxl.utils.exceptions as ox_exc

from ImportWindow import Ui_Import_Target
from definitions import GeneralMessageExt

import messageFunctions


class ExcelInterface:
    """
    Interface class contains functions to interface between the PLC object and the output file.
    """

    def __init__(self, plc=None, filepath=''):
        assert isinstance(plc, pylogix.PLC)
        self.plc = plc  # pointer to plc object
        self._filePath = ''
        self.workbook = None
        self.filePath = filepath
        self.progDict = {}

    @property
    def filePath(self):
        return self._filePath

    @filePath.setter
    def filePath(self, path):
        # Filepath is set by save dialog, checking validity is redundant
        self._filePath = path

    def Export(self):
        try:
            # Create Workbook
            self.workbook = xlsxwriter.Workbook(self._filePath)
            sheetNames = self.findPrograms()
            # create sheet names
            for sheetName in sheetNames:  # Loop Programs as sheets
                progName = str.replace(sheetName, 'Program:', '')
                worksheet = self.workbook.add_worksheet(progName)
                bold = self.workbook.add_format({'bold': True})  # create bold formatting
                wrap = self.workbook.add_format({'text_wrap': True})
                faultTags = messageFunctions.loadFaults(self.plc, progName)
                messageTags = messageFunctions.loadMessages(self.plc, progName)
                # set column width
                worksheet.set_column(0, 0, 10)  # Format ID Columns
                worksheet.set_column(4, 4, 10)  # Format ID Columns
                worksheet.set_column(1, 2, 40)  # Format Text Columns
                worksheet.set_column(5, 6, 40)  # Format Text Columns
                worksheet.set_column(3, 3, 15)  # Format Divider
                # Build out Headers
                worksheet.write('A1', 'Fault ID', bold)
                worksheet.write('B1', 'Text', bold)
                worksheet.write('C1', 'AltText', bold)
                # Skip Column D For readability
                worksheet.write('E1', 'Message ID', bold)
                worksheet.write('F1', 'Text', bold)
                worksheet.write('G1', 'AltText', bold)
                # enumerate and write out tag data for Faults
                for index, tag in enumerate(faultTags):
                    index = index + 2  # Avoid overwriting the Headers
                    id = worksheet.write(f'A{index}', tag.Id, wrap)
                    text = worksheet.write(f'B{index}', tag.Text, wrap)
                    altText = worksheet.write(f'C{index}', tag.AltText, wrap)
                    if id or text or altText:
                        raise xlsx_exc.XlsxWriterException(
                            f'Writer Fault Status: Id:{id}, Text:{text}, AltText:{altText}')
                # enumerate this for messages since they may have different lengths
                for index, tag in enumerate(messageTags):
                    index = index + 2  # Avoid overwriting the Headers
                    id = worksheet.write(f'E{index}', tag.Id, wrap)
                    text = worksheet.write(f'F{index}', tag.Text, wrap)
                    altText = worksheet.write(f'G{index}', tag.AltText, wrap)
                    if id or text or altText:
                        raise xlsx_exc.XlsxWriterException(
                            f'Writer Message Status: Id:{id}, Text:{text}, AltText:{altText}')
            self.workbook.close()

        except xlsx_exc.FileCreateError:
            msg = QMessageBox(QMessageBox.Warning, 'File Create Error',
                              'XlSX writer cannot edit file.\n'
                              'Ensure the file is not open by another process and try again.',
                              QMessageBox.Ok)
            msg.exec()

        except xlsx_exc.XlsxWriterException:
            msg = QMessageBox(QMessageBox.Warning, 'Errors Detected',
                              'XLSX writer returned an error during the write process.\n '
                              'The output file may not be accurate.',
                              QMessageBox.Ok)
            msg.exec()
        else:
            msg = QMessageBox(QMessageBox.Information, 'Complete', 'Tag Export Complete.', QMessageBox.Ok)
            msg.exec()

    def Import(self):
        """
        Import function uses openpyxl library to open and import data for read operations.
        :return:
        """
        try:
            self.workbook = openpyxl.load_workbook(self._filePath, read_only=True)
            progNames = self.workbook.sheetnames  # return list of sheet names in excel
            progressBox = QProgressDialog('Doing super duper stuff...', '', 0, len(progNames))
            progressBox.setWindowModality(Qt.WindowModal)
            progressBox.show()

            for index, prog in enumerate(progNames):
                progressBox.setLabelText(f'Importing tgs from "{prog}"')
                sheet = self.workbook[prog]
                if sheet.max_column != 7:
                    raise Exception('Invalid Excel input format')
                rowCount = sheet.max_row
                read = sheet[f'A2:G{rowCount}']
                """
                sheet read object returns a tuple of rows
                To access individual cells, example: 'read[0][0].value' returns the value of the above 'read' variable 
                as row[2]col[A]. Remember excel rows start at 1 while the tuples will start at 0. 
                """
                faultMessages = []
                operatorMessages = []
                progressBox.setValue(index)
                for row in read:  # iterate read rows and access by column index
                    try:
                        fm = GeneralMessageExt()
                        Id = row[0].value
                        if Id is not None:
                            fm.newId = row[0].value  # Fault Id
                            fm.newText = row[1].value  # Fault Text
                            fm.newAltText = row[2].value  # Fault AltText
                            faultMessages.append(fm)
                    except IndexError:
                        # Index errors will occur as message arrays will not match length, simply ignore
                        pass
                    try:
                        Id = row[4].value
                        if Id is not None:
                            om = GeneralMessageExt()
                            om.newId = row[4].value  # Message ID
                            om.newText = row[5].value   # Message Text
                            om.newAltText = row[5].value  # Message AltText
                            operatorMessages.append(om)
                    except IndexError:
                        # Index errors will occur as message arrays will not match length, simply ignore
                        pass
                    # Add Fault and Message tags to a dictionary as a tuple
                    tags = (faultMessages, operatorMessages)
                    self.progDict[prog] = tags
            progressBox.setValue(len(progNames))  # Should kill the progress dialog box
            progressBox.close()

            window = Import_Window(progNames)
            if window.exec():  # execute the window and wait for the return

                for prog in window.targetList:
                    messageFunctions.send_faults(plc=self.plc, progName=prog, tagList=self.progDict[prog][0])
                    messageFunctions.send_messages(plc=self.plc, progName=prog, tagList=self.progDict[prog][1])

        except ox_exc.InvalidFileException as e:
            msg = QMessageBox(QMessageBox.Warning, 'Invalid File Exception',
                              'Exception was thrown when attempting top read from file.\n'
                              f'{e}',
                              QMessageBox.Ok)
            msg.exec()

        except ox_exc.NamedRangeException as e:
            msg = QMessageBox(QMessageBox.Warning, 'Invalid Range Exception',
                              'Named Range Exception during Import.\n'
                              f'{e}',
                              QMessageBox.Ok)
            msg.exec()

    def findPrograms(self):
        """
        Returns a list of Programs Names that contain valid message fault and operator tags.
        :return: List[str]
        """
        programList = []
        attempts = 2
        while attempts > 0:
            # Get The list of programs from the PLC
            ret = self.plc.GetProgramsList()
            attempts -= 1
            # check return for success
            if ret.Status == 'Success':
                programList = ret.Value
                break  # exit while
            else:
                if attempts == 0:
                    # TODO: Look into returning a Message Message Dialog box
                    return -1  # Failed by communication

        # Look for 'MessageArrayFault' and MessageArrayOperator and remove programs without those tags
        filteredProgList = []
        for prog in programList:
            target1 = prog + '.MessageArrayFault'
            target2 = prog + '.MessageArrayOperator'
            for tag in self.plc.TagList:
                if tag.TagName == target1 or tag.TagName == target2:
                    filteredProgList.append(prog)
                    break
        if len(programList) != len(filteredProgList):
            print(f'Programs List was reduced from {len(programList)} to {len(filteredProgList)}')
        # Filter the 'Program:' from the front of the names to use as sheet names
        print(filteredProgList)
        return filteredProgList


class Import_Window(Ui_Import_Target, QDialog):
    def __init__(self, progList):
        super(Import_Window, self).__init__()
        self.setupUi(self)
        self.setWindowTitle('Tag Import Selection')
        self.pbn_write.clicked.connect(self.action_writeTags)
        self.targetList = []
        self.progList = progList
        self.model = QStandardItemModel()
        self.buildCheckList()

    def buildCheckList(self):
        # Build Program list in listbox out of checkboxes
        for prog in self.progList:
            item = QStandardItem(prog)
            item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            item.setData(QVariant(Qt.Checked), Qt.CheckStateRole)
            self.model.appendRow(item)
        self.lst_chkboxes.setModel(self.model)

    def action_writeTags(self):
        # Get program list of checked programs and pass to tag writer
        for index in range(self.model.rowCount()):
            item = self.model.item(index, 0)  # row by index, column 0 since only one column
            if item.checkState():
                self.targetList.append(item.text())

        self.accept()  # Closes dialogbox
